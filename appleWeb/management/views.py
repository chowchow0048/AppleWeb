from django.shortcuts import render, get_object_or_404, redirect
from django.views.decorators.http import require_POST, require_GET
from django.contrib.auth.decorators import login_required
from django.contrib.auth import get_user_model
from django.contrib import messages
from django.core.paginator import Paginator
from django.core.exceptions import ObjectDoesNotExist
from django.http import JsonResponse, HttpResponse
from django.utils.dateparse import parse_date
from django.utils import timezone
from django.db import transaction
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from common.models import (
    User,
    Course,
    Attendance,
    Absence,
    Handover,
    Waitlist,
    Blacklist,
)
from common.decorators import manager_required
import pandas as pd
import logging
import urllib.parse
import traceback


logger_appleWeb = logging.getLogger("appleWeb")


@login_required
@manager_required
def api_students(request):
    logger_appleWeb.debug("AWDB: Request received with params: %s", request.GET)

    school = request.GET.get("school")
    grade = request.GET.get("grade")
    students_query = User.objects.filter(is_active=True)

    if school:
        students_query = students_query.filter(school=school)
    if grade and grade != "전체":
        students_query = students_query.filter(grade=grade)

    students_data = sorted(
        [
            {
                "id": student.id,
                "school": student.school,
                "grade": student.grade,
                "name": student.name,
                "phone": student.phone,
                "parent_phone": student.parent_phone,
                "payment_count": student.payment_count,
            }
            for student in students_query
        ],
        key=lambda student: student["name"],  # name을 기준으로 오름차순 정렬
    )

    return JsonResponse(students_data, safe=False)


@login_required
@manager_required
def api_courses(request):
    logger_appleWeb.debug("AWDB: Request received with params: %s", request.GET)

    day = request.GET.get("day")
    school = request.GET.get("school")

    courses = Course.objects.filter(course_school=school, is_active=True)

    if day:
        courses = courses.filter(course_day=day)

    data = [
        {
            "id": course.id,
            "course_grade": course.course_grade,
            "course_subject": course.course_subject,
            "course_time": course.course_time.strftime("%H:%M"),
            "course_day": course.course_day,
        }
        for course in courses
    ]
    return JsonResponse(data, safe=False)


@login_required
@manager_required
def export_attendance_to_excel(request, course_id):
    course = Course.objects.get(id=course_id)
    students = course.course_students.all()  # 코스에 등록된 모든 학생들을 불러옴

    # 엑셀 파일 제목 줄 설정
    today = timezone.now().strftime("%Y-%m-%d")
    subject = course.course_subject
    if subject == "physics":
        subject = "물리"
    elif subject == "chemistry":
        subject = "화학"
    elif subject == "biology":
        subject = "생명과학"
    elif subject == "earth_science":
        subject = "지구과학"
    else:
        subject = "통합과학"

    title = f"{course.course_school} {course.course_grade} {subject} / {course.course_day} {course.course_time.strftime('%H:%M')} / {today}"
    filename = f"{course.course_school}{course.course_grade}_{subject}_{course.course_day}_{course.course_time.strftime('%H')}시_출석부.xlsx"
    encoded_filename = urllib.parse.quote(filename.encode("utf-8"))

    # 워크북 생성
    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance"
    ws.append(["학교", "학년", "이름", "전화번호", "부모님 전화번호", "출석"])

    # 학생 데이터 추가
    for student in students:
        ws.append(
            [
                student.school,
                student.grade,
                student.name,
                student.phone if student.phone else "-",
                student.parent_phone if student.parent_phone else "-",
                "",  # 출석
            ]
        )

    # 열 너비 조정
    for col in range(1, 8):  # 열 A부터 G까지
        ws.column_dimensions[get_column_letter(col)].width = 15

    # 제목 행 병합 및 스타일 설정
    ws.insert_rows(1)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)
    ws.cell(row=1, column=1).value = title
    ws.cell(row=1, column=1).alignment = Alignment(horizontal="center")

    # HTTP 응답 설정
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    # UTF-8로 인코딩된 파일 이름 적용
    response["Content-Disposition"] = f"attachment; filename*=UTF-8''{encoded_filename}"

    # 엑셀 파일 저장 및 전송
    wb.save(response)
    return response


@login_required
@manager_required
def management_home(request):
    return render(request, "management/management_home.html")


@login_required
@manager_required
def management_all_lectures(request):
    return render(request, "management/management_all_lectures.html")


@login_required
@manager_required
def management_studentlist(request):
    return render(request, "management/management_studentlist.html")


@login_required
@manager_required
def management_lecture(request, course_id):
    course = Course.objects.get(id=course_id)
    students = course.course_students.all().order_by("name")
    today = timezone.now().date()

    # print("course:", course)
    # print("students:", students)
    # print("course_id:", course_id)
    # print("today:", today)

    attendance_records = {
        attendance.student.id: attendance
        for attendance in Attendance.objects.filter(course=course, date=today)
    }

    absence_records = {
        absence.student.id: absence
        for absence in Absence.objects.filter(course=course, date=today)
    }

    return render(
        request,
        "management/management_lecture.html",
        {
            "course": course,
            "students": students,
            "attendance_records": attendance_records,
            "absence_records": absence_records,
        },
    )


@login_required
@manager_required
@require_POST  # 이 뷰 함수는 POST 요청만 허용
def bulk_attendance(request):
    date = timezone.now().date()  # 오늘 날짜를 가져옴
    course_id = request.POST.get("course_id")  # POST 데이터에서 course_id를 가져옴

    if not course_id:
        return redirect("management_home")  # course_id가 없으면 홈 페이지로 리디렉션

    course = get_object_or_404(
        Course, id=course_id
    )  # 주어진 course_id에 해당하는 Course 객체를 가져옴

    attendance_ids = request.POST.getlist(
        "attendance"
    )  # POST 데이터에서 attendance 리스트를 가져옴
    absence_ids = request.POST.getlist(
        "absence"
    )  # POST 데이터에서 absence 리스트를 가져옴

    # 기존 출석 및 결석 기록을 초기화하기 전에 기존 출석 기록의 결제 횟수를 복구
    old_attendance_records = Attendance.objects.filter(course=course, date=date)
    old_absence_records = Absence.objects.filter(course=course, date=date)

    # print("OLD ATTENDANCE RECORD:", old_attendance_records)
    # print("OLD ABSENCE RECORD:", old_absence_records)

    # 기존 출석 기록을 삭제하고, 결제 횟수를 복구합니다
    for record in old_attendance_records:
        student = record.student  # 학생 객체를 가져옴
        student.payment_count += 1  # 결제 횟수를 복구
        student.save()

    for record in old_absence_records:
        student = record.student
        student.payment_count += 1
        student.save()

    old_attendance_records.delete()  # 기존 출석 기록 삭제
    old_absence_records.delete()  # 기존 결석 기록 삭제

    # 새로운 출석 기록을 생성하고, 결제 횟수를 차감
    for student_id in attendance_ids:
        student = course.course_students.get(
            id=student_id
        )  # student_id에 해당하는 학생을 가져옴
        Attendance.objects.create(
            course=course, student=student, date=date
        )  # 새로운 출석 기록 생성
        student.payment_count -= 1  # 결제 횟수를 차감
        if student.payment_count <= 0:
            student.payment_request = True

        student.save()

    # 새로운 결석 기록을 생성
    for student_id in absence_ids:
        student = course.course_students.get(
            id=student_id
        )  # student_id에 해당하는 학생을 가져옴
        Absence.objects.create(
            course=course, student=student, date=date
        )  # 새로운 결석 기록 생성
        student.payment_count -= 1  # 결제 횟수를 차감
        if student.payment_count <= 0:
            student.payment_request = True

        student.save()

    return redirect(
        "management_lecture", course_id=course_id
    )  # 출석부 페이지로 리디렉션


@login_required
@manager_required
def management_student_detail(request, student_id):
    user = request.user
    student = get_object_or_404(User, pk=student_id)
    attendances = Attendance.objects.filter(student=student).order_by("attended_at")
    absences = Absence.objects.filter(student=student).order_by("absent_at")
    courses = student.enrolled_courses.filter(is_active=True).all()

    if not (user.is_superuser or user.is_manager or user.id == student_id):
        messages.error(request, "접근 권한이 없습니다.")
        return redirect("main/403.html")  # 접근 권한이 없을 때 리디렉션할 뷰

    if request.method == "POST":
        password1 = request.POST.get("password1")
        password2 = request.POST.get("password2")
        if password1 and password2 and password1 == password2:
            student.set_password(password1)
            student.save()
            messages.success(request, "비밀번호가 성공적으로 변경되었습니다.")
        else:
            messages.error(request, "비밀번호가 일치하지 않습니다.")

    context = {
        "student": student,
        "attendances": attendances,
        "absences": absences,
        "courses": courses,
    }

    return render(request, "management/management_student_detail.html", context)


@login_required
@manager_required
def management_handover(request):
    handovers = Handover.objects.all().order_by("-created_date")
    paginator = Paginator(handovers, 10)
    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)

    context = {"page_obj": page_obj}

    return render(request, "management/management_handover.html", context)


# 인수인계 검색 보류
# @login_required
# @manager_required
# def search_handover(request):
#     query = request.GET.get("query", "")
#     results = []

#     try:
#         print(f"Search query: {query}")  # 검색어 출력

#         if query:
#             handovers = Handover.objects.filter(title__icontains=query).order_by(
#                 "-created_date"
#             )
#             for handover in handovers:
#                 results.append(
#                     {
#                         "id": handover.id,
#                         "title": f"{handover.created_date.strftime('%Y-%m-%d')} {handover.shift} 행정",
#                         "author": handover.author.username,
#                         "created_date": handover.created_date.strftime("%Y-%m-%d"),
#                     }
#                 )

#         print(f"Results: {results}")  # 결과 출력
#         return JsonResponse({"results": results})

#     except Exception as e:
#         print(f"Error: {e}")
#         print(traceback.format_exc())  # 전체 오류 스택을 로그로 출력
#         return JsonResponse({"error": "Something went wrong"}, status=500)


@login_required
@manager_required
def management_handover_detail(request, handover_id):
    handover = get_object_or_404(Handover, id=handover_id)
    return render(
        request, "management/management_handover_detail.html", {"handover": handover}
    )


@login_required
@manager_required
def management_handover_delete(request, handover_id):
    handover = get_object_or_404(Handover, id=handover_id)
    if handover.author == request.user:
        handover.delete()
        return redirect("management_handover")  # 삭제 후 목록 페이지로 리다이렉트
    return redirect("management_handover_detail", handover_id=handover_id)


@login_required
@manager_required
def management_handover_add(request):
    today = timezone.now().strftime("%Y-%m-%d")

    if request.method == "POST":
        # POST 데이터에서 shift와 content를 직접 가져옵니다.
        shift = request.POST.get("shift")
        content = request.POST.get("content")

        if shift and content:
            # Handover 객체를 생성하여 데이터베이스에 저장
            handover = Handover(
                author=request.user,  # 작성자 설정
                shift=shift,
                content=content,
                created_date=timezone.now(),  # 현재 날짜 설정
            )
            handover.save()

            # 성공적으로 저장 후 리다이렉트 처리 (원하는 경로로 수정)
            return redirect("management_handover")

    context = {"today": today}
    return render(request, "management/management_handover_add.html", context)


@login_required
@manager_required
def management_handover_update(request, handover_id):
    handover = get_object_or_404(Handover, id=handover_id)

    # 작성자가 아닌 경우 접근 제한
    if handover.author != request.user:
        return redirect("management_handover_detail", handover_id=handover_id)

    if request.method == "POST":
        # POST 요청에서 데이터를 수동으로 받아옴
        shift = request.POST.get("shift")
        content = request.POST.get("content")

        # 데이터가 유효한 경우 업데이트
        if shift and content:
            handover.shift = shift
            handover.content = content
            handover.save()
            return redirect("management_handover_detail", handover_id=handover_id)

    context = {
        "handover": handover,
        "today": handover.created_date.strftime("%Y-%m-%d"),  # 기존 날짜를 그대로 사용
    }

    return render(request, "management/management_handover_update.html", context)


@login_required
@manager_required
def management_paylist(request):
    users_with_payment_request = User.objects.filter(
        payment_request=True, is_active=True
    )
    users_with_payment_request = [
        {
            "id": user.id,
            "school": user.school,
            "grade": user.grade,
            "name": user.name,
            "phone": user.phone,
            "parent_phone": user.parent_phone,
            "payment_count": abs(user.payment_count),
        }
        for user in users_with_payment_request
    ]
    print(users_with_payment_request)

    return render(
        request,
        "management/management_paylist.html",
        {"users": users_with_payment_request},
    )


@login_required
@manager_required
@require_GET  # 이 뷰는 GET 요청만 허용
def fetch_paylist(request):
    school = request.GET.get("school", None)
    query = request.GET.get("query", "")

    users = User.objects.filter(payment_request=True, is_active=True, is_teacher=False)

    # 학교 필터 적용
    if school:
        users = users.filter(school=school)

    # 이름 검색 필터 적용
    if query:
        users = users.filter(name__icontains=query)

    # 필요한 필드만 JSON으로 반환
    data = list(
        users.values(
            "id", "school", "grade", "name", "phone", "parent_phone", "payment_count"
        )
    )
    return JsonResponse(data, safe=False)


@login_required
@manager_required
@require_POST  # 이 뷰 함수는 POST 요청만 허용
def confirm_payment(request, user_id):
    user = get_object_or_404(User, pk=user_id)

    # 수강하고있는 강의 수 만큼 결제횟수 갱신
    course_count = user.enrolled_courses.filter(is_active=True).count()
    user.payment_count += course_count * 4  # 기존 payment_count에 course 수 x 4를 더함
    user.payment_request = False
    user.latest_payment = timezone.now().date()  # 최근 결제 날짜를 오늘 날짜로 설정
    user.save()

    return JsonResponse(
        {
            "message": "결제 요청 완료",
            "payment_count": user.payment_count,
        }
    )


@login_required
@manager_required
def management_wait_black_list(request):
    return render(request, "management/management_wait_black_list.html")


@login_required
@manager_required
def fetch_wait_black_list(request):
    list_type = request.GET.get("type", "wait")
    school = request.GET.get("school", None)
    query = request.GET.get("query", None)
    print(list_type)
    if list_type == "wait":
        queryset = Waitlist.objects.all()
    else:
        queryset = Blacklist.objects.all()

    if school:
        queryset = queryset.filter(school=school)

    if query:
        queryset = queryset.filter(name__icontains=query)

    # 데이터를 list로 변환, id와 list_type을 포함
    data = list(
        queryset.values("id", "school", "grade", "name", "phone", "date", "note")
    )
    for entry in data:
        entry["list_type"] = list_type  # 각 엔트리에 list_type 추가

    return JsonResponse(data, safe=False)


@login_required
@manager_required
def management_wait_black_list_add(request, list_type):
    today = timezone.now().strftime("%Y-%m-%d")

    if request.method == "POST":
        school = request.POST.get("school")
        grade = request.POST.get("grade")
        name = request.POST.get("name")
        phone = request.POST.get("phone")
        note = request.POST.get("note")

        if list_type == "wait":
            Waitlist.objects.create(
                school=school,
                grade=grade,
                name=name,
                phone=phone,
                note=note,
                author=request.user,
            )
        else:
            Blacklist.objects.create(
                school=school,
                grade=grade,
                name=name,
                phone=phone,
                note=note,
                author=request.user,
            )

        return redirect("management_wait_black_list")

    context = {"today": today, "list_type": list_type}
    return render(
        request, f"management/management_wait_black_list_{list_type}_add.html", context
    )


@login_required
@manager_required
def management_wait_black_list_delete(request, list_type, entry_id):
    # 'list_type'은 'wait' 또는 'black'을 받아 대기/블랙리스트 구분
    if list_type == "wait":
        entry = get_object_or_404(Waitlist, id=entry_id)
    else:
        entry = get_object_or_404(Blacklist, id=entry_id)

    entry.delete()  # 항목 삭제
    return redirect("management_wait_black_list")


@login_required
@manager_required
def management_wait_black_list_wait_detail(request, waitlist_id):
    wait = get_object_or_404(Waitlist, id=waitlist_id)

    context = {
        "wait": wait,
        "school": wait.school,
        "grade": wait.grade,
        "name": wait.name,
        "phone": wait.phone,
        "note": wait.note,
        "date": wait.date,
        "author": wait.author,
    }

    return render(
        request, "management/management_wait_black_list_wait_detail.html", context
    )


@login_required
@manager_required
def management_wait_black_list_black_detail(request, blacklist_id):
    black = get_object_or_404(Blacklist, id=blacklist_id)

    context = {
        "black": black,
        "school": black.school,
        "grade": black.grade,
        "name": black.name,
        "phone": black.phone,
        "note": black.note,
        "date": black.date,
        "author": black.author,
    }

    return render(
        request, "management/management_wait_black_list_black_detail.html", context
    )
